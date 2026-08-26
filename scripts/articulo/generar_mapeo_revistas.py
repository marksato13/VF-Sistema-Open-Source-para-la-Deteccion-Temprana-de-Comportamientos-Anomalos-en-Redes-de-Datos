#!/usr/bin/env python3
"""Mapeo por secciones de 10 articulos de BEEI e IJSSE afines al PPI.

Las secciones, parrafos, citas, tablas y figuras se extrajeron leyendo el PDF
completo de cada articulo, no de resumenes ni de metadatos.

    python3 scripts/articulo/generar_mapeo_revistas.py
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[2] / "docs/articulo/mapeo-secciones-BEEI-IJSSE.xlsx"

AZUL = "1F3864"; CLARO = "D9E2F3"; CREMA = "FFF2CC"; GRIS = "F2F2F2"
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)

BEEI = [
 ("Hybrid AI-driven anomaly detection and sequential attack classification for securing IoT networks",
  "10.11591/eei.v15i1.11048", "2026 · Vol. 15, No. 1 · pp. 669-679 · 30 referencias",
  ["1. INTRODUCTION -- Solo 3 párrafos pero con ~23 citas, la sección más densa en referencias de todo el artículo. Plantea el crecimiento del IoT y su superficie de ataque, repasa en bloque los enfoques previos (ML clásico, DL, híbridos) y cierra enunciando las contribuciones. No usa subsecciones.",
   "2. METHOD -- La sección más extensa: 23 párrafos, 4 subsecciones y prácticamente sin citas (1), porque describe lo propio. Subsecciones: 2.1 Dataset description · 2.2 Preprocessing and feature selection · 2.3 Anomaly detection · 2.4 Intrusion classification. Incluye la Tabla 1 y las Figuras 1-3 (arquitectura del marco de dos niveles: RF binario como filtro rápido y LSTM multiclase).",
   "3. RESULT -- 13 párrafos con 6 citas y una fuerte carga visual: Tablas 2, 3, 4 y 5 más la Figura 4. Dos subsecciones: 3.1 Outcome discussion y, notablemente, 3.2 Scope and limitation, donde declara las limitaciones dentro de la sección de resultados en vez de esconderlas.",
   "4. CONCLUSION -- 5 párrafos sin ninguna cita. Resume el 99 % de exactitud sobre NF-ToN-IoT-V2 y los 7,8 ms de respuesta, y proyecta trabajo futuro."]),
 ("Detecting anomalies in MQTT/MQTT-SN traffic using intelligent learning models",
  "10.11591/eei.v15i3.11889", "2026 · Vol. 15, No. 3 · pp. 2758-2771 · 30 referencias",
  ["1. INTRODUCTION -- Muy breve: 2 párrafos con 13 citas. Presenta MQTT/MQTT-SN como protocolos IoT dominantes y su falta de mecanismos de seguridad nativos.",
   "2. MATERIALS -- 3 párrafos con 18 citas. Funciona como estado del arte comprimido: en vez de una sección 'Related Work' aparte, concentra aquí la revisión y el material de partida.",
   "3. PROPOSED MODEL -- 22 párrafos y 10 subsecciones, sin citas: todo es descripción propia. 3.1 Data collection and annotation · 3.2 Feature extraction and engineering · 3.3 Preprocessing · 3.4 Model design and algorithm implementation (3.4.2 Random forest, 3.4.3 LSTM) · 3.5 Model training and validation (3.5.1 Performance metrics, 3.5.2 Deployment and resource profiling). Tablas 1-2 y Figuras 1-2.",
   "4. RESULTS AND DISCUSSION -- 19 párrafos, 5 citas, Tablas 3 a 7 y 7 subsecciones: 4.1 Classification performance · 4.2 Confusion matrix analysis · 4.3 MQTT vs. MQTT-SN detection performance · 4.4 Execution time and resource usage · 4.5 Discussion and implications · 4.6 Generalization to real-world IoT environments · 4.7 Ethical and legal considerations. Las dos últimas son inusuales y valiosas: discuten generalización fuera del laboratorio y aspectos éticos.",
   "5. CONCLUSION -- 9 párrafos, la conclusión más larga de las cinco de BEEI. Sin citas."]),
 ("An efficient intrusion detection systems in fog computing using forward selection and BiLSTM",
  "10.11591/eei.v13i4.7143", "2024 · Vol. 13, No. 4 · pp. 2586-2603 · 54 referencias",
  ["1. INTRODUCTION -- 2 párrafos con 8 citas. Muy escueta; plantea el problema y salta rápido al desarrollo.",
   "2. BACKGROUND -- 8 párrafos y 19 citas, con la Figura 1 y 4 subsecciones: 2.1 Fog computing · 2.2 Fog security issues · 2.3 Intrusion detection system · 2.4 Features selection. Separa el marco conceptual del estado del arte, que va aparte.",
   "3. RELATED WORK -- 2 párrafos con 8 citas. Independiente del background: aquí solo se contrastan trabajos previos.",
   "4. METHOD -- La sección dominante: 27 párrafos, 12 citas, Figuras 2-6 y 11 subsecciones. Cubre preprocesamiento (4.1.1 codificación, 4.1.2 limpieza y valores faltantes, 4.1.3 normalización) y el método propio de selección de variables con su función hiperobjetivo (4.2.1), información mutua (4.2.2), correlación de Pearson (4.2.3), varianza (4.2.4), entropía (4.2.5) y el clasificador BiLSTM (4.2.6).",
   "5. EXPERIMENTAL SETUP -- 11 párrafos, 3 citas, Tabla 1 y 5 subsecciones. Sección propia para el montaje experimental, separada de los resultados.",
   "6. RESULTS -- 16 párrafos, 4 citas, Tablas 2-4 y Figuras 7 a 12, con 4 subsecciones.",
   "7. CONCLUSION -- 2 párrafos, sin citas."]),
 ("Anomaly intrusion detection using machine learning - IG-R based on NSL-KDD dataset",
  "10.11591/eei.v13i6.7308", "2024 · Vol. 13, No. 6 · pp. 4466-4474 · 26 referencias",
  ["1. INTRODUCTION -- 2 párrafos con 10 citas.",
   "2. RELATED WORK -- Un solo párrafo con 7 citas. Es la sección de estado del arte más breve de las diez analizadas.",
   "3. DATASET, ADVERSARY MODEL, AND CONFUSION MATRIX -- 8 párrafos con Tablas 1 y 2. Estructura poco común: agrupa en una misma sección el dataset, el modelo de adversario y la definición de la matriz de confusión, antes de presentar el método.",
   "4. TESTING MODEL -- 23 párrafos, 5 citas, Figuras 1-5 y 6 subsecciones: 4.1 Data preprocessing and instances selection · 4.2 Data classification, con una subsección por clasificador (4.2.1 AdaBoost · 4.2.2 Random forest · 4.2.3 J48 · 4.2.4 Naïve Bayes).",
   "5. RESULTS EVALUATION AND ANALYSIS -- Solo 4 párrafos, sin citas, con la Tabla 3 y la Figura 6. Muy comprimida frente al desarrollo del método.",
   "6. CONCLUSION -- 1 párrafo."]),
 ("A stacked ensemble approach to identify internet of things network attacks through traffic analysis",
  "10.11591/eei.v13i6.7811", "2024 · Vol. 13, No. 6 · pp. 4316-4326 · 43 referencias",
  ["1. INTRODUCTION -- 3 párrafos con 16 citas.",
   "2. RELATED WORK -- 6 párrafos y 22 citas, con la Tabla 1 y 3 subsecciones: 2.1 Internet of things security landscape · 2.2 Anomaly detection for IoT security · 2.3 IoTID20 dataset. Dedica una subsección al dataset dentro del estado del arte.",
   "3. METHOD -- 9 párrafos, 10 citas, Figura 1 y 7 subsecciones: 3.1 System overview · 3.2 IoTID20 dataset preprocessing · 3.3 Feature selection · 3.4 Training data balancing · 3.5 Base classifiers · 3.6 Meta-classifier · 3.7 Evaluation metrics. Una subsección por componente del ensamble.",
   "4. EXPERIMENTS AND RESULTS -- 7 párrafos con Tablas 2-4 y Figura 2, organizados por tipo de tarea: 4.1 Binary classification · 4.2 Multi-class classification · 4.3 Multi-label classification.",
   "5. DISCUSSION -- 12 párrafos, 6 citas y Tablas 5 a 8. Sección separada de los resultados, con 5.1 Comparison of ensemble versus individual classifiers y 5.2 Statistical validation of results: incluye validación estadística explícita, algo poco frecuente en este corpus.",
   "6. CONCLUSION -- 1 párrafo."]),
]

IJSSE = [
 ("FL-NDR: A Federated Unlearning-Driven Network Detection and Response System for DDoS Defence in Software-Defined Networks",
  "10.18280/ijsse.160615", "2026 · Vol. 16, No. 6 · pp. 1333-1342 · 25 referencias",
  ["1. INTRODUCTION -- 4 párrafos con ~9 citas. Plantea la vulnerabilidad del controlador SDN ante DDoS y enuncia las contribuciones del sistema.",
   "2. RELATED WORK -- 3 párrafos con 4 subsecciones temáticas: 2.1 DDoS attack detection in SDN · 2.2 Network detection and response · 2.3 Federated Learning for network security · 2.4 Adversarial attacks and unlearning. Cada subsección cierra nombrando la limitación del trabajo previo y cómo la cubre la propuesta.",
   "3. FEDERATED UNLEARNING NDR SYSTEM DESIGN -- 9 párrafos, Tabla 1 y Figuras 1-2, con 4 subsecciones: 3.1 SDN fabric as a distributed sensor layer · 3.2 Four layer architecture · 3.3 Response Orchestration Layer · 3.4 Federated Unlearning Pipeline. Describe respuesta en tres niveles (T1 limitación de tasa, T2 descarte, T3 cuarentena de dominio).",
   "4. SDN SIMULATION VALIDATION -- 6 párrafos con la Tabla 2 y Figuras 3-4, en 2 subsecciones: 4.1 Sensor telemetry signatures · 4.2 Response Orchestration Layer behaviour. Sección propia para validar el comportamiento del sistema antes de evaluar el modelo.",
   "5. FEDERATED UNLEARNING EVALUATION -- 9 párrafos, Tablas 2-4 y Figuras 5-6, con 4 subsecciones: 5.1 Dataset · 5.2 Threat model and configuration · 5.3 Experimental scenarios · 5.4 Results and analysis. Compara seis escenarios (S1-S6).",
   "6. DISCUSSIONS -- 5 párrafos con 13 citas y 3 subsecciones de título argumentativo: 6.1 Quarantine is not optional · 6.2 Why the sensor fabric matters · 6.3 Limitations. Titula las subsecciones con la conclusión, no con el tema.",
   "7. CONCLUSIONS -- 1 párrafo."]),
 ("Hybrid Machine Learning-Based Intrusion Detection for Zero-Day Attack Prevention in Digital Education Networks",
  "10.18280/ijsse.150815", "2025 · Vol. 15, No. 8 · pp. 1703-1713 · 27 referencias",
  ["1. INTRODUCTION -- 4 párrafos con 8 citas. Abre con un dato de impacto (el sector educativo como el más atacado) y expone el límite de los IDS por firmas frente a los ataques de día cero.",
   "2. RELATED WORK -- 13 párrafos y 24 citas, la revisión más extensa de las diez, con 8 subsecciones: 2.1 Signature-based · 2.2 Anomaly-based · 2.3 Machine learning-based · 2.4 Hybrid and ensemble (2.4.1 Hybrid signature-anomaly · 2.4.2 Ensemble ML · 2.4.3 Feature-hybrid methods · 2.4.4 Drawbacks of existing hybrid methods). La última subsección enuncia explícitamente el vacío que el artículo llena.",
   "3. PROPOSED METHODOLOGY AND THEORETICAL FRAMEWORK -- 8 párrafos, Tabla 1 y Figuras 1-3, con 4 subsecciones: 3.1 Autoencoder anomaly detector · 3.2 Random Forest classifier · 3.3 Combining the modules · 3.4 Model training and computational complexity. El autocodificador modela el tráfico normal y marca lo desviado: mismo principio no supervisado que el PPI.",
   "4. COMPARATIVE ANALYSIS WITH EXISTING TECHNIQUES -- 4 párrafos con la Tabla 2 y Figuras 4-5, en 2 subsecciones. Sección propia para el protocolo de comparación y las líneas base, antes de mostrar resultados.",
   "5. RESULTS ON UNIVERSITY NETWORK DATASET -- 8 párrafos con 14 citas y 3 subsecciones, una de ellas con cinco sub-subsecciones (5.3.1 zero-day detection · 5.3.2 low false positive rate · 5.3.3 robustness · 5.3.4 applicability · 5.3.5 implications). Reporta 99,1 % de exactitud y menos de 1 % de falso positivo.",
   "6. CONCLUSIONS -- Muy breve, sin citas."]),
 ("Deep Reinforcement Learning-Based Energy-Aware Intrusion Prevention in IoT Environment",
  "10.18280/ijsse.150819", "2025 · Vol. 15, No. 8 · pp. 1745-1754 · 22 referencias",
  ["1. INTRODUCTION -- 3 párrafos con 12 citas. Cierra enunciando tres contribuciones numeradas.",
   "2. RELATED WORK -- 2 párrafos con 10 citas, sin subsecciones. Muy comprimida.",
   "3. PROPOSED MODEL -- 21 párrafos sin ninguna cita, con la Figura 1 y 6 subsecciones: 3.1 System model and assumptions · 3.2 DRL-EAIPS (3.2.1 State representation · 3.2.2 Action space · 3.2.3 Reward function · 3.2.4 Policy learning). Carga matemática alta: ecuaciones 1 a 14.",
   "4. RESULTS AND DISCUSSION -- 12 párrafos con 11 citas, Tablas 1-4 y Figuras 2 a 6. Fusiona resultados y discusión en una sola sección y evalúa tres despliegues (50, 100 y 150 nodos) sobre tres datasets.",
   "5. CONCLUSIONS -- Breve, sin citas."]),
 ("A Comparative Study of Incremental and Batch Machine Learning Methodologies for Network Intrusion Detection",
  "10.18280/ijsse.150118", "2025 · Vol. 15, No. 1 · pp. 171-179 · 25 referencias",
  ["1. INTRODUCTION -- 4 párrafos con 16 citas. Identifica el vacío: falta comparación sistemática entre aprendizaje incremental y por lotes en IDS.",
   "2. BACKGROUND -- 5 párrafos, 10 citas, Figuras 1-3 y 4 subsecciones: 2.1 Intrusion detection system · 2.2 Machine learning (2.2.1 Incremental · 2.2.2 Batch) · 2.3 Literature review. La revisión bibliográfica va dentro del background, no como sección propia.",
   "3. PROPOSED METHODOLOGY -- 9 párrafos, 9 citas, Figura 4 y 4 subsecciones: 3.1 Dataset (UNSW-NB15 y CIC-IDS2017) · 3.2 Data preparation, con cinco etapas · 3.3 Model design · 3.4 Evaluation, con las ecuaciones 2 a 8.",
   "4. IMPLEMENTATION AND EXPERIMENTAL RESULTS -- 3 párrafos con Tablas 1-2 y Figuras 5-6 (matrices de confusión), en 5 subsecciones: 4.1 y 4.2 resultados por dataset · 4.3 Implications and contributions · 4.4 Comparison with existing literature · 4.5 Practical applications.",
   "5. DISCUSSION -- 5 párrafos con 14 citas. Sección aparte de los resultados.",
   "6. CONCLUSIONS -- 1 párrafo."]),
 ("Deep Learning Based Intrusion Detection System of IoT Technology: Accuracy Versus Computational Complexity",
  "10.18280/ijsse.140522", "2024 · Vol. 14, No. 5 · pp. 1547-1558 · 34 referencias",
  ["1. INTRODUCTION -- 5 párrafos con 17 citas y la Figura 1 (diagrama de bloques del IDS). Plantea explícitamente el compromiso entre exactitud y coste computacional, que es el eje del artículo.",
   "2. RELATED WORK -- 3 párrafos con 14 citas y la Tabla 1, que compara dataset, clasificador y exactitud de los trabajos previos en una sola vista.",
   "3. THE PROPOSED APPROACH'S DESIGN -- 15 párrafos, 9 citas, Figuras 2 a 8 y 8 subsecciones: 3.1 LSTM (3.1.1 forget gate · 3.1.2 input gate · 3.1.3 output gate) · 3.2 Bi-LSTM · 3.3 GRU · 3.4 GRU con mecanismo de atención · 3.5 Grid Search. Ecuaciones 1 a 20.",
   "4. DATASET DESCRIPTION AND COLLECTION -- Sección propia y corta para el dataset ToN-IoT (461 043 registros, 45 columnas), con la Tabla 2.",
   "5. EVALUATION METRICS FOR MODEL PERFORMANCE -- Sección propia solo para definir métricas, con 5 subsecciones: 5.1 Confusion matrix · 5.2 Accuracy · 5.3 Precision · 5.4 Recall · 5.5 F1-score.",
   "6. RESULT AND DISCUSSION -- Tablas 3 a 9 y Figuras 9 a 15 (curvas de exactitud y pérdida, matrices de confusión), con la subsección 6.1 sobre los resultados con Grid Search. Reporta 99 % de exactitud con 0,3 ms de clasificación.",
   "7. CONCLUSION -- Destaca la reducción del 84 % en tiempo de clasificación."]),
]


def hoja(wb, nombre, datos, revista, url):
    ws = wb.create_sheet(nombre)
    maxs = max(len(d[3]) for d in datos)
    ws["A1"] = f"{revista} — {url}"
    ws["A1"].font = Font(bold=True, size=12, color=AZUL)
    heads = ["N°", "Título", "DOI", "Metadatos"] + [f"Sección {i+1}" for i in range(maxs)]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDE
    for r, (tit, doi, meta, secs) in enumerate(datos, start=3):
        fill = PatternFill("solid", fgColor=GRIS if r % 2 else "FFFFFF")
        vals = [r - 2, tit, doi, meta] + secs + [""] * (maxs - len(secs))
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDE
            cell.fill = fill
            if c == 2:
                cell.font = Font(bold=True)
            if c == 3:
                cell.font = Font(color="0563C1")
        ws.row_dimensions[r].height = 190
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 30
    for i in range(5, 5 + maxs):
        ws.column_dimensions[get_column_letter(i)].width = 58
    ws.freeze_panes = "E3"
    return ws


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Cómo se hizo")
    filas = [
        ("Mapeo por secciones — 10 artículos afines al PPI", ""),
        ("", ""),
        ("Objetivo", "Ver cómo estructuran sus artículos las dos revistas objetivo, para escribir el propio en su misma forma."),
        ("Revista 1", "Bulletin of Electrical Engineering and Informatics (BEEI) — https://beei.org/index.php/EEI"),
        ("Revista 2", "International Journal of Safety and Security Engineering (IJSSE) — https://www.iieta.org/Journals/IJSSE"),
        ("Criterio de selección", "Afinidad con el PPI: detección de anomalías o intrusiones en red mediante aprendizaje automático. Se priorizó 2026 > 2025 > 2024."),
        ("Cómo se obtuvieron los datos", "Se descargó el PDF completo de cada artículo y se extrajo su texto. Las secciones, párrafos, citas, tablas y figuras están contados sobre el texto real, no tomados de resúmenes."),
        ("Precisión de las cifras", "Los conteos de párrafos son aproximados: se cuenta cada bloque de texto separado por línea en blanco con más de 120 caracteres. Las citas son referencias distintas [n] dentro de la sección."),
        ("", ""),
        ("Aviso", "Los cinco ejemplos del archivo original son revisiones sistemáticas (RSL). Estos diez son artículos EMPÍRICOS, que es lo que pide el sílabo del curso como producto académico."),
    ]
    for r, (a, b) in enumerate(filas, 1):
        ws.cell(row=r, column=1, value=a).font = Font(bold=True, color=AZUL, size=13 if r == 1 else 11)
        c = ws.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110
    for r in range(3, 11):
        ws.row_dimensions[r].height = 34

    hoja(wb, "BEEI", BEEI, "Bulletin of Electrical Engineering and Informatics",
         "https://beei.org/index.php/EEI")
    hoja(wb, "IJSSE", IJSSE, "International Journal of Safety and Security Engineering",
         "https://www.iieta.org/Journals/IJSSE")

    ws = wb.create_sheet("Patrones")
    pat = [
        ("Patrón observado", "BEEI (5 artículos)", "IJSSE (5 artículos)"),
        ("Número de secciones", "4 a 7 · mediana 6", "5 a 7 · mediana 6"),
        ("Sección dominante", "METHOD / PROPOSED MODEL — hasta 27 párrafos", "PROPOSED MODEL / METHODOLOGY — hasta 21 párrafos"),
        ("Dónde se concentran las citas", "Introduction y Related Work; el método casi no cita", "Related Work; en IJSSE la revisión llega a 24 citas"),
        ("Resultados y discusión", "3 de 5 los separan en secciones distintas", "3 de 5 los fusionan en una sola sección"),
        ("Conclusión", "De 1 a 9 párrafos, casi siempre sin citas", "Breve, 1 párrafo, sin citas"),
        ("Total de referencias", "26 a 54", "22 a 34"),
        ("Subsecciones", "Hasta 11 en el método", "Hasta 8 en el estado del arte"),
        ("Sección propia para el dataset", "No: va dentro del método", "Sí en 2 de 5, como sección numerada aparte"),
        ("Declaración de limitaciones", "1 de 5 (subsección dentro de Resultados)", "2 de 5 (subsección dentro de Discusión)"),
    ]
    for r, fila in enumerate(pat, 1):
        for c, v in enumerate(fila, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDE
            if r == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=AZUL)
            elif c == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=CREMA)
        ws.row_dimensions[r].height = 32
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 52

    wb.save(OUT)
    print(f"Generado: {OUT}")
    print(f"  hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
