#!/usr/bin/env python3
"""Entregable 3: mapeo de artículos y estructura final del artículo.

Cumple lo pedido en clase: mínimo 12 documentos (aquí 21), con al menos 6
artículos semilla de la revista principal (aquí 7 de IJIES), los nombres de
sección **tal cual aparecen impresos**, la estadística de frecuencias, el
veredicto de estructura rígida o flexible, el análisis crítico de la guía de
autores y la estructura final que se deriva de las dos fuentes.

    .venv/bin/python3 scripts/articulo/generar_mapeo_estructura.py
"""
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

SALIDA = "docs/entregables/10-mapeo-secciones-articulo/Mapeo-estructura-articulo.xlsx"

AZUL, GRIS, VERDE, AMBAR, ROJO, BLANCO = (
    "1F3864", "F2F2F2", "E2EFDA", "FFF2CC", "FCE4EC", "FFFFFF")
BORDE = Border(*[Side("thin", color="BFBFBF")] * 4)

# ---------------------------------------------------------------------------
# 21 artículos. Los nombres de sección se transcriben EXACTAMENTE como los
# imprime cada artículo, incluidas mayúsculas y erratas del original.
# (revista, nº, título, DOI, año/vol/pp, págs., refs., [secciones], [finales])
# ---------------------------------------------------------------------------
ART = [
 # ---- 7 SEMILLA · IJIES (revista principal, Opción 1) --------------------
 ("IJIES", 1, "HIDE-6G: Advanced Intrusion Detection System for Secure 6G Network "
  "using Deep Learning", "10.22266/ijies2024.1031.37", "2024 · 17(5) · 474-483", 10, 17,
  ["1. Introduction", "2. Literature survey",
   "3. Hybrid intrusion detection system for 6G", "4. Results and discussion",
   "5. Conclusion"],
  ["Conflicts of Interest", "Author Contributions", "Acknowledgments", "References"]),
 ("IJIES", 2, "Optimizing Intrusion Detection in Internet of Things (IoT) Networks "
  "Using a Hybrid PSO-LightBoost Approach", "10.22266/ijies2025.0430.14",
  "2025 · 18(3) · 195-208", 14, 22,
  ["1. Introduction", "2. Literature review", "3. Proposed methodology",
   "4. Result analysis & discussion", "5. Conclusion with future work"],
  ["Conflicts of Interest", "Author Contributions", "References"]),
 ("IJIES", 3, "Optimizing Feature Selection Method in Intrusion Detection System "
  "Using Thresholding", "10.22266/ijies2024.0630.18", "2024 · 17(3) · 214-226", 13, 44,
  ["1. Introduction", "2. Related work", "3. Methodology",
   "4. Result and discussion", "5. Conclusion"],
  ["Conflicts of Interest", "Author Contributions", "Acknowledgments (con funding)",
   "References"]),
 ("IJIES", 4, "Efficient Two-Stage Intrusion Detection System Based on Hybrid Feature "
  "Selection Techniques and Machine Learning", "10.22266/ijies2025.0430.16",
  "2025 · 18(3) · 224-240", 17, 39,
  ["1. Introduction", "2. Related works", "3. Attack detection in WSNs",
   "4. Intrusion detection architecture for IoT",
   "5. Hybrid intrusion detection with cuckoo search", "6. Results and discussion",
   "7. Conclusion"],
  ["Conflicts of Interest", "Author Contributions", "References"]),
 ("IJIES", 5, "Analysis of Weight-Based Voting Classifier for Intrusion Detection System",
  "10.22266/ijies2024.0430.17", "2024 · 17(2) · 190-200", 11, 39,
  ["1. Introduction", "2. Related works", "3. The proposed method",
   "4. Results and discussion", "5. Conclusion"],
  ["Conflicts of interest", "Author contributions", "Acknowledgments", "References"]),
 ("IJIES", 6, "Network Intrusion Detection System Based on Information Gain with Deep "
  "Bidirectional Long Short-Term Memory", "10.22266/ijies2024.0831.04",
  "2024 · 17(4) · 45-56", 12, 26,
  ["1. Introduction", "2. Literature survey", "3. Methodology",
   "4. Results and discussion", "5. Conclusion"],
  ["Conflicts of Interest", "Author Contributions", "References"]),
 ("IJIES", 7, "Network Intrusion Detection Using Feature Selection Techniques: "
  "Bacterial Forage Optimization Algorithm", "10.22266/ijies2024.1031.48",
  "2024 · 17(5) · 630-645", 16, 49,
  ["1. Introduction", "2. Related work", "3. System methodology",
   "4. Evaluation metrics", "5. Results and discussion", "6. Conclusion"],
  ["Conflicts of Interest", "Author Contributions", "References"]),
 # ---- 4 · ISI (Opción 2) --------------------------------------------------
 ("ISI", 8, "Statistical Anomaly Detection for Enhanced Cybersecurity Using AI-Based "
  "Wireless Networks", "10.18280/isi.290508", "2024 · 29(5) · 1743-1754", 12, 40,
  ["1. INTRODUCTION", "2. LITERATURE REVIEW", "3. SUPER VECTOR MACHINE (SVM)",
   "4. METHODOLOGY", "5. RESULTS", "6. DISCUSSION", "7. CONCLUSION"], ["REFERENCES"]),
 ("ISI", 9, "Machine Learning-Based Anomaly Detection Model for Cybersecurity Threat "
  "Detection", "10.18280/isi.290628", "2024 · 29(6) · 2415-2424", 10, 22,
  ["1. INTRODUCTION", "2. RELATED WORK", "3. METHODOLOGY",
   "4. REUSLTS AND DISCUSSION  [errata del original]", "5. CONCLUSION"], ["REFERENCES"]),
 ("ISI", 10, "Machine Learning for Cloud Data Classification and Anomaly Intrusion "
  "Detection", "10.18280/isi.290514", "2024 · 29(5) · 1809-1819", 11, 41,
  ["1. INTRODUCTION", "2. RELATED WORKS", "3. BACKGROUND", "4. PROPOSED METHODOLOGY",
   "5. EXPERIMENTAL RESULT", "6. CONCLUSIONS"], ["REFERENCES"]),
 ("ISI", 11, "Deep Learning Based Multistage Approach for Anomaly Detection",
  "10.18280/isi.290534", "2024 · 29(5) · 2031-2038", 8, 22,
  ["1. INTRODUCTION", "2. RELATED WORK", "3. THE PRESENT METHODOLOGY",
   "4. EXPERIMENTATION AND RESULTS", "5. PERFORMANCE ANALYSIS",
   "6. CONCLUSION AND FUTURE SCOPE"], ["REFERENCES", "ACKNOWLEDGEMENT"]),
 # ---- 5 · BEEI (Opción 3) -------------------------------------------------
 ("BEEI", 12, "Hybrid AI-driven anomaly detection and sequential attack classification "
  "for securing IoT networks", "10.11591/eei.v15i1.11048", "2026 · 15(1) · 669-679",
  11, 30, ["1. INTRODUCTION", "2. METHOD", "3. RESULT", "4. CONCLUSION"], ["REFERENCES"]),
 ("BEEI", 13, "Detecting anomalies in MQTT/MQTT-SN traffic using intelligent learning "
  "models", "10.11591/eei.v15i3.11889", "2026 · 15(3) · 2758-2771", 14, 30,
  ["1. INTRODUCTION", "2. MATERIALS", "3. PROPOSED MODEL",
   "4. RESULTS AND DISCUSSION", "5. CONCLUSION"], ["REFERENCES"]),
 ("BEEI", 14, "An efficient intrusion detection systems in fog computing using forward "
  "selection and BiLSTM", "10.11591/eei.v13i4.7143", "2024 · 13(4) · 2586-2603", 18, 54,
  ["1. INTRODUCTION", "2. BACKGROUND", "3. RELATED WORK", "4. METHOD",
   "5. EXPERIMENTAL SETUP", "6. RESULTS", "7. CONCLUSION"], ["REFERENCES"]),
 ("BEEI", 15, "Anomaly intrusion detection using machine learning - IG-R based on "
  "NSL-KDD dataset", "10.11591/eei.v13i6.7308", "2024 · 13(6) · 4466-4474", 9, 26,
  ["1. INTRODUCTION", "2. RELATED WORK",
   "3. DATASET, ADVERSARY MODEL, AND CONFUSION MATRIX", "4. TESTING MODEL",
   "5. RESULTS EVALUATION AND ANALYSIS", "6. CONCLUSION"], ["REFERENCES"]),
 ("BEEI", 16, "A stacked ensemble approach to identify internet of things network "
  "attacks through traffic analysis", "10.11591/eei.v13i6.7811",
  "2024 · 13(6) · 4316-4326", 11, 43,
  ["1. INTRODUCTION", "2. RELATED WORK", "3. METHOD", "4. EXPERIMENTS AND RESULTS",
   "5. DISCUSSION", "6. CONCLUSION"], ["REFERENCES"]),
 # ---- 5 · IJSSE (reserva) -------------------------------------------------
 ("IJSSE", 17, "FL-NDR: A Federated Unlearning-Driven Network Detection and Response "
  "System", "10.18280/ijsse.160615", "2026 · 16(6) · 1333-1342", 10, 25,
  ["1. INTRODUCTION", "2. RELATED WORK",
   "3. FEDERATED UNLEARNING NDR SYSTEM DESIGN", "4. SDN SIMULATION VALIDATION",
   "5. FEDERATED UNLEARNING EVALUATION", "6. DISCUSSIONS", "7. CONCLUSIONS"],
  ["REFERENCES"]),
 ("IJSSE", 18, "Hybrid Machine Learning-Based Intrusion Detection for University "
  "Networks", "10.18280/ijsse.150815", "2025 · 15(8) · 1703-1713", 11, 27,
  ["1. INTRODUCTION", "2. RELATED WORK",
   "3. PROPOSED METHODOLOGY AND THEORETICAL FRAMEWORK",
   "4. COMPARATIVE ANALYSIS WITH EXISTING TECHNIQUES",
   "5. RESULTS ON UNIVERSITY NETWORK DATASET", "6. CONCLUSIONS"], ["REFERENCES"]),
 ("IJSSE", 19, "Deep Reinforcement Learning-Based Energy-Aware Intrusion Detection",
  "10.18280/ijsse.150819", "2025 · 15(8) · 1745-1754", 10, 22,
  ["1. INTRODUCTION", "2. RELATED WORK", "3. PROPOSED MODEL",
   "4. RESULTS AND DISCUSSION", "5. CONCLUSIONS"], ["REFERENCES"]),
 ("IJSSE", 20, "A Comparative Study of Incremental and Batch Machine Learning",
  "10.18280/ijsse.150118", "2025 · 15(1) · 171-179", 9, 25,
  ["1. INTRODUCTION", "2. BACKGROUND", "3. PROPOSED METHODOLOGY",
   "4. IMPLEMENTATION AND EXPERIMENTAL RESULTS", "5. DISCUSSION", "6. CONCLUSIONS"],
  ["REFERENCES"]),
 ("IJSSE", 21, "Deep Learning Based Intrusion Detection System of IoT",
  "10.18280/ijsse.140522", "2024 · 14(5) · 1547-1558", 12, 34,
  ["1. INTRODUCTION", "2. RELATED WORK", "3. THE PROPOSED APPROACH'S DESIGN",
   "4. DATASET DESCRIPTION AND COLLECTION",
   "5. EVALUATION METRICS FOR MODEL PERFORMANCE", "6. RESULT AND DISCUSSION",
   "7. CONCLUSION"], ["REFERENCES"]),
]

# Familia funcional de cada nombre literal. Se declara a mano y a la vista para
# que cualquiera pueda discutir la agrupación; el nombre literal nunca se pierde.
FAMILIA = [
 ("Introducción", ["introduction"]),
 ("Estado del arte", ["related work", "related works", "literature review",
                      "literature survey", "materials"]),
 ("Marco teórico previo", ["background", "super vector machine (svm)",
                           "attack detection in wsns",
                           "intrusion detection architecture for iot"]),
 ("Metodología / Modelo propuesto",
  ["methodology", "proposed methodology", "the present methodology",
   "system methodology", "method", "proposed model", "the proposed method",
   "hybrid intrusion detection system for 6g",
   "hybrid intrusion detection with cuckoo search",
   "proposed methodology and theoretical framework",
   "the proposed approach's design", "testing model",
   "federated unlearning ndr system design",
   "dataset, adversary model, and confusion matrix",
   "dataset description and collection", "experimental setup"]),
 ("Métricas de evaluación", ["evaluation metrics",
                             "evaluation metrics for model performance"]),
 ("Resultados", ["results", "result", "experimental result",
                 "experimentation and results", "experiments and results",
                 "performance analysis", "results evaluation and analysis",
                 "results on university network dataset", "sdn simulation validation",
                 "federated unlearning evaluation",
                 "comparative analysis with existing techniques",
                 "implementation and experimental results"]),
 ("Resultados y discusión (fusionados)",
  ["results and discussion", "result and discussion", "result analysis & discussion",
   "reusits and discussion", "reuslts and discussion"]),
 ("Discusión", ["discussion", "discussions"]),
 ("Conclusión", ["conclusion", "conclusions", "conclusion with future work",
                 "conclusion and future scope"]),
]

GUIA = [
 ("¿Impone una estructura de secciones?", "**NO.** La guía no nombra ni una sola "
  "sección del cuerpo del artículo.",
  "La estructura debe justificarse con el mapeo. Es el caso previsto: la guía no la "
  "manifiesta, los artículos publicados sí la determinan."),
 ("Extensión", "«**Standard number of pages is 8. (The papers must be 8 pages or "
  "more.)**»", "Es un **mínimo de 8**, no un máximo. El artículo no puede ser corto."),
 ("Sobrecosto por extensión", "«Extra page charge **USD 50 per extra page** will be "
  "requested if the paper length **exceeds 10 pages**.»",
  "La banda sin sobrecosto es **8 a 10 páginas**. Los 7 semilla miden 10, 11, 12, 13, "
  "14, 16 y 17: **cinco de siete pagaron extra.** Apuntar a 10."),
 ("Plantilla", "`IJIES_Format.docx` obligatoria · «extra fee **USD 100** if authors do "
  "not use the IJIES_Format.docx»", "Descargar y escribir dentro de la plantilla desde "
  "el primer borrador."),
 ("Tipo de artículo", "«The review article (**Survey research**) **is not acceptable** "
  "for publication.»", "El nuestro es experimental. Cumple."),
 ("Proceso de revisión", "Dos rondas (1st y 2nd review), «**about 2 weeks**» cada una; "
  "dos revisores; notificación en 2 semanas. Decisiones **A / B / C / D**.",
  "Coincide con lo medido: mediana de **41 días** a la primera decisión sobre 5 "
  "artículos."),
 ("Publicación", "Mensual · «Accepted papers will be published **about 2 months "
  "later**»", "Mediana medida de recepción a publicación: **158 días** (≈5 meses)."),
 ("**Declaración de uso de IA**", "Política propia: «**Transparency regarding "
  "substantial AI use is required**». Permite corrección de idioma, depuración de "
  "código, organización de ideas y preparación de figuras. **Prohíbe** listar la IA "
  "como autor y entregar contenido sin verificación humana.",
  "**Obligatorio declararlo**: el proyecto usó asistentes de IA. Va como sección "
  "propia al final."),
 ("Cesión de derechos", "«upon acceptance, the **copyright of the article will be "
  "transferred to the publisher**» · formulario *Copyright Transfer* firmado.",
  "Explica que su acceso abierto sea **bronce sin licencia**: se lee gratis pero el "
  "autor cede el copyright."),
 ("Plagio", "«a fine of **USD 5,000** is requested to authors» si se detecta plagio.",
  "Pasar iThenticate o similar antes de enviar."),
 ("Carta de presentación", "«Effective **May 19, 2025**, the submission of cover "
  "letters has been rendered **unnecessary**»", "No preparar carta. Solo el manuscrito."),
 ("Pago", "«**Paypal (Credit card) is the only payment method**. The bank transfer is "
  "**not supported**.»", "Resolver el medio de pago **antes** de enviar."),
 ("APC", "USD 300 hoy · **USD 400 desde el 1 de octubre de 2026** (incluye ya el cargo "
  "de formato)", "Enviar antes de octubre ahorra USD 100."),
]

ESTRUCTURA = [
 ("1", "Introduction", "21/21 — **100 %**", "Guía: no la nombra · Mapeo: unánime",
  "Contexto, brecha, aporte y contribuciones. Aquí va el argumento de novedad: "
  "ninguno de los 21 cierra el lazo detección → bloqueo automático."),
 ("2", "Related work", "13/21 con ese nombre · **19/21 (90 %)** como familia",
  "Guía: no la nombra · Mapeo: nombre dominante", "Trabajos previos de detección de "
  "anomalías en red. Es donde se concentran las citas."),
 ("3", "Proposed methodology", "**21/21 — 100 %** como familia",
  "Guía: no la nombra · Mapeo: presente en los 21",
  "Laboratorio, dataset multilayer-v2, las 28 variables, ventanas, OCSVM y "
  "calibración del umbral. Sección más extensa (hasta 36 párrafos en la muestra). "
  "El dataset va **aquí como subsección**, no como sección aparte (así en 3 de 4 "
  "revistas)."),
 ("4", "Results and discussion", "**11/21 fusionados** · 10/21 separados",
  "Guía: no la nombra · Mapeo: **fusionar es lo más frecuente en IJIES (5/7)**",
  "Métricas offline, validación operativa F6 y la brecha 4,71 % → 23-26 %. "
  "Incluye la subsección de limitaciones."),
 ("5", "Conclusion", "21/21 — **100 %**", "Guía: no la nombra · Mapeo: unánime",
  "Sin citas: **0 de 21 artículos citan en la conclusión**. Cierre y trabajo futuro."),
 ("6", "Conflicts of Interest", "7/7 en IJIES — **100 %**",
  "Guía: no la exige · Mapeo IJIES: unánime", "Declaración breve."),
 ("7", "Author Contributions", "7/7 en IJIES — **100 %**",
  "Guía: no la exige · Mapeo IJIES: unánime",
  "Reparto entre Salazar Tocas y Sauñe Fernandez."),
 ("8", "Declaración de uso de IA", "0/21 en la muestra",
  "**Guía: obligatoria** («transparency regarding substantial AI use is required») · "
  "Mapeo: aún no aparece",
  "**La exige la guía aunque no esté en los artículos de 2024.** Qué asistió la IA "
  "y qué verificaron los autores."),
 ("9", "Acknowledgments", "3/7 en IJIES — 43 %",
  "Guía: no la exige · Mapeo: opcional",
  "Asesores y financiamiento, si lo hay. Uno de los siete mete el *funding* aquí."),
 ("10", "References", "21/21 — **100 %**", "Guía: no fija estilo · Mapeo: unánime",
  "Rango observado en IJIES: **17 a 49**, mediana 39."),
]


def estilar(ws, ancho):
    for i, a in enumerate(ancho, 1):
        ws.column_dimensions[get_column_letter(i)].width = a
    for fila in ws.iter_rows():
        for c in fila:
            c.border = BORDE
            c.alignment = Alignment(wrap_text=True, vertical="top")


def titulo(ws, txt, ncol, color=AZUL):
    ws.append([txt] + [""] * (ncol - 1))
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row,
                   end_column=ncol)
    c = ws.cell(ws.max_row, 1)
    c.font = Font(bold=True, size=12, color=BLANCO)
    c.fill = PatternFill("solid", fgColor=color)
    ws.row_dimensions[ws.max_row].height = 26


def nota(ws, txt, ncol, alto=32):
    ws.append([txt] + [""] * (ncol - 1))
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row,
                   end_column=ncol)
    ws.cell(ws.max_row, 1).font = Font(italic=True, size=9)
    ws.row_dimensions[ws.max_row].height = alto


def cabecera(ws, cols, color=GRIS):
    ws.append(cols)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=color)


def familia_de(nombre):
    n = nombre.split(". ", 1)[-1].strip().lower()
    n = n.split("  [errata")[0].strip()
    for fam, claves in FAMILIA:
        if n in claves:
            return fam
    return None


def main():
    wb = Workbook()

    # ------------------------------------------------------------ 1. mapeo --
    ws = wb.active
    ws.title = "1. MAPEO"
    maxs = max(len(a[7]) for a in ART)
    ncol = 7 + maxs + 1
    titulo(ws, "Mapeo de 21 artículos — nombres de sección transcritos tal cual los "
               "imprime cada artículo", ncol)
    nota(ws, "7 artículos semilla de IJIES (revista Opción 1) y 14 de otras revistas. "
             "Mínimo exigido: 12 documentos, 6 de ellos semilla. Cada fila lleva su DOI "
             "para verificar. Las erratas del original se conservan y se señalan.", ncol)
    cabecera(ws, ["Revista", "N°", "Título", "DOI", "Publicación", "Págs.", "Refs."] +
                 [f"Sección {i+1}" for i in range(maxs)] +
                 ["Otras secciones (tras las referencias)"])
    for rev, n, tit, doi, pub, pags, refs, secs, fin in ART:
        ws.append([rev, n, tit, doi, pub, pags, refs] + secs +
                  [""] * (maxs - len(secs)) + [" · ".join(fin)])
        ws.cell(ws.max_row, 1).fill = PatternFill(
            "solid", fgColor=VERDE if rev == "IJIES" else GRIS)
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=9)
        ws.cell(ws.max_row, 3).font = Font(bold=True, size=9)
    estilar(ws, [9, 4, 40, 24, 20, 7, 7] + [27] * maxs + [34])
    ws.freeze_panes = "C4"

    # ------------------------------------------------------ 2. frecuencias --
    lit = Counter()
    litci = Counter()
    fam = Counter()
    for *_, secs, _f in ART:
        for s in secs:
            nom = s.split(". ", 1)[-1].split("  [errata")[0].strip()
            lit[nom] += 1
            litci[nom.lower()] += 1
        for f in sorted({familia_de(s) for s in secs if familia_de(s)}):
            fam[f] += 1
    total = len(ART)

    ws = wb.create_sheet("2. FRECUENCIAS")
    titulo(ws, "Frecuencia de secciones sobre los 21 artículos", 4)
    nota(ws, "Tres lecturas. (1) El nombre literal, tal cual está impreso — lo que pide "
             "la consigna. Ojo: IJIES escribe «Introduction» y las demás «INTRODUCTION», "
             "así que el mismo nombre aparece partido; es estilo de la casa editorial, no "
             "una sección distinta. (2) El mismo nombre ignorando mayúsculas. (3) La "
             "familia funcional, que es la que permite decidir la estructura: «Related "
             "work», «Literature review» y «Literature survey» son la misma sección con "
             "tres nombres.", 4, 62)
    cabecera(ws, ["Nombre literal, tal cual aparece impreso", "Artículos", "% de 21",
                  "Familia funcional"])
    for nom, c in sorted(lit.items(), key=lambda x: (-x[1], x[0])):
        f = familia_de(nom) or "—"
        ws.append([nom, c, round(100 * c / total, 1), f])
        if c >= total * 0.5:
            ws.cell(ws.max_row, 2).font = Font(bold=True, size=9)
    ws.append([])
    cabecera(ws, ["Mismo nombre, ignorando mayúsculas", "Artículos", "% de 21",
                  "Familia funcional"], VERDE)
    for nom, c in sorted(litci.items(), key=lambda x: (-x[1], x[0])):
        if c < 2:
            continue
        ws.append([nom, c, round(100 * c / total, 1), familia_de(nom) or "—"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=9)
    fila_fam = ws.max_row + 2
    ws.append([])
    cabecera(ws, ["Familia funcional", "Artículos", "% de 21", ""], AMBAR)
    for f, c in sorted(fam.items(), key=lambda x: (-x[1], x[0])):
        ws.append([f, c, round(100 * c / total, 1), ""])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=9)
    ini_fam = fila_fam + 1
    ws.append([])
    ws.append([f"Nombres literales distintos: {len(lit)}", "",
               f"Familias funcionales: {len(fam)}", ""])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=9)
    ws.cell(ws.max_row, 3).font = Font(bold=True, size=9)

    graf = BarChart()
    graf.type = "bar"
    graf.title = "Familias de sección presentes (de 21 artículos)"
    graf.y_axis.title = None
    graf.height, graf.width = 9, 17
    graf.add_data(Reference(ws, min_col=2, min_row=ini_fam,
                            max_row=ini_fam + len(fam) - 1), titles_from_data=False)
    graf.set_categories(Reference(ws, min_col=1, min_row=ini_fam,
                                  max_row=ini_fam + len(fam) - 1))
    graf.legend = None
    ws.add_chart(graf, f"F{fila_fam}")
    estilar(ws, [48, 11, 10, 34])
    ws.freeze_panes = "A4"

    # --------------------------------------------------- 3. rígido/flexible --
    ws = wb.create_sheet("3. RIGIDO O FLEXIBLE")
    titulo(ws, "¿La estructura de la revista es rígida o flexible?", 3)
    intro_pct = round(100 * lit["Introduction"] / total)
    ws_rows = [
     ("Criterio", "Medición sobre los 21 artículos", "Lectura"),
    ]
    cabecera(ws, list(ws_rows[0]))
    for c, m, l in [
     ("Nombres literales distintos",
      f"**{len(lit)} nombres distintos** para {sum(lit.values())} secciones de cuerpo",
      "Muchísima variación léxica"),
     ("Secciones unánimes",
      "Solo **2 de 21**: «Introduction» (100 %) y la familia «Conclusión» (100 %)",
      "El principio y el final están fijados; el medio no"),
     ("Nombre del estado del arte",
      "**5 nombres** para la misma función: Related work · Related works · Literature "
      "review · Literature survey · Materials",
      "Nadie impone el nombre"),
     ("Nombre de la metodología",
      "**16 nombres distintos**, desde «Methodology» hasta «Hybrid intrusion detection "
      "system for 6G»", "Se acepta incluso titularla con el nombre del método propio"),
     ("Número de secciones de cuerpo",
      "Entre **4 y 7** · mediana 5 en IJIES, 6 en el resto",
      "No hay número fijo"),
     ("Resultados y discusión",
      "9 de 21 los fusionan, 10 los separan, 2 usan otro arreglo",
      "Ni siquiera esto está estandarizado"),
     ("Secciones tras las referencias",
      "En IJIES: **Conflicts of Interest 7/7 y Author Contributions 7/7**; "
      "Acknowledgments 3/7", "**Aquí sí hay rigidez**, y es la parte que suele olvidarse"),
     ("Lo que dice la guía de autores",
      "**No nombra ni una sola sección del cuerpo**",
      "No hay imposición desde la norma"),
    ]:
        ws.append([c, m, l])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=9)
    ws.append([])
    titulo(ws, "VEREDICTO: estructura FLEXIBLE en el cuerpo, RÍGIDA en el cierre", 3,
           VERDE)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, size=12, color="1F3864")
    nota(ws, f"IJIES no impone secciones y sus artículos usan {len(lit)} nombres "
             "distintos, así que podemos elegir los nuestros. Lo que sí es obligatorio: "
             "abrir con «Introduction» (100 %), cerrar con «Conclusion» (100 %) y "
             "cerrar con Conflicts of Interest y Author Contributions (7/7 en IJIES). "
             "Elegimos los nombres más frecuentes, no porque haya que hacerlo, sino "
             "porque un revisor los reconoce sin esfuerzo.", 3, 60)
    estilar(ws, [30, 62, 44])

    # -------------------------------------------------- 4. guía de autores --
    ws = wb.create_sheet("4. GUIA DE AUTORES")
    titulo(ws, "Análisis crítico de la guía de autores de IJIES (Opción 1)", 3)
    nota(ws, "Fuentes leídas el 02/09/2026: inass.org/pub-submissionguidelines/ · "
             "inass.org/pub-charges/ · inass.org/pub-docusubmission/. "
             "Las citas entre comillas son literales.", 3)
    cabecera(ws, ["Aspecto", "Lo que dice la guía, textual", "Qué implica para nosotros"])
    for a, b, c in GUIA:
        ws.append([a, b, c])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=9)
    ws.append([])
    nota(ws, "CONCLUSIÓN DEL ANÁLISIS · La guía de IJIES regula el formato, el pago, la "
             "ética y el uso de IA, pero NO la estructura: no menciona ninguna sección "
             "del cuerpo. Por eso la estructura final se justifica con el mapeo de los "
             "artículos publicados. Dos exigencias de la guía no aparecen en los "
             "artículos de 2024 y hay que añadirlas igual: la declaración de uso de IA "
             "y el mínimo de 8 páginas.", 3, 62)
    estilar(ws, [30, 62, 50])
    ws.freeze_panes = "A4"

    # ------------------------------------------------- 5. estructura final --
    ws = wb.create_sheet("5. ESTRUCTURA FINAL")
    titulo(ws, "Estructura final del artículo — justificada por las dos fuentes", 5)
    nota(ws, "Cada sección se sostiene en el mapeo de los 21 artículos Y en la guía de "
             "autores. Una vez aprobada, es la estructura con la que se redacta.", 5)
    cabecera(ws, ["N°", "Sección", "Frecuencia en el mapeo",
                  "Respaldo de las dos fuentes", "Qué va dentro"])
    for n, s, fr, j, q in ESTRUCTURA:
        ws.append([n, s, fr, j, q])
        ws.cell(ws.max_row, 2).font = Font(bold=True, size=10)
        ws.cell(ws.max_row, 2).fill = PatternFill("solid", fgColor=VERDE)
    ws.append([])
    nota(ws, "OTRAS SECCIONES (conflictos de interés, contribución de autores, "
             "declaración de uso de inteligencia artificial, agradecimientos y "
             "financiamiento si lo hubiera, y referencias). Son las que van después del "
             "cuerpo del artículo: en IJIES, Conflicts of Interest y Author "
             "Contributions aparecen en los 7 de 7 artículos revisados, y la "
             "declaración de IA la exige la guía aunque todavía no aparezca en los "
             "artículos de 2024.", 5, 62)
    ws.append([])
    ws.append(["Extensión objetivo: 10 páginas — mínimo 8 exigido por la guía, y desde "
               "la 11.ª cuesta USD 50 por página.", "", "", "", ""])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=10)
    ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=AMBAR)
    ws.row_dimensions[ws.max_row].height = 24
    estilar(ws, [5, 26, 24, 40, 56])
    ws.freeze_panes = "A4"

    # ------------------------------------------------------------ 6. ficha --
    ws = wb.create_sheet("0. COMO SE HIZO")
    titulo(ws, "Mapeo y estructura del artículo — cómo se construyó", 2)
    cabecera(ws, ["Aspecto", "Detalle"])
    for a, b in [
     ("Equipo", "Rubén Mark Salazar Tocas · Uziel Elias Sauñe Fernandez"),
     ("Fecha", "2 de septiembre de 2026"),
     ("Revista Opción 1", "IJIES — International Journal of Intelligent Engineering and "
      "Systems (INASS, Japón) · Q2 · CiteScore 3,3 · APC USD 300"),
     ("Revista Opción 2", "ISI — Ingénierie des Systèmes d'Information (IIETA) · Q3 · "
      "APC USD 850"),
     ("Revista Opción 3", "BEEI — Bulletin of Electrical Engineering and Informatics "
      "(IAES) · Q1 por CiteScore · APC USD 415"),
     ("Documentos analizados", f"**{len(ART)}**, por encima del mínimo de 12: "
      "**7 semilla de IJIES** (la Opción 1) y 14 de otras revistas — 4 de ISI, 5 de "
      "BEEI y 5 de IJSSE"),
     ("Criterio de selección", "Afinidad con el tema: detección de intrusiones o de "
      "anomalías en redes, publicados entre 2024 y 2026"),
     ("Cómo se leyeron", "Se descargó el PDF completo de cada artículo desde su DOI y se "
      "transcribieron los encabezados **tal cual están impresos**, sin normalizar ni "
      "traducir. Las erratas del original se conservan y se señalan entre corchetes"),
     ("Por qué el PDF y no el DOI", "La página del DOI muestra a veces una lista de "
      "secciones que no coincide con el artículo. La única fuente fiable del nombre real "
      "es el PDF"),
     ("Secciones tras las referencias", "Se registraron aparte, porque la consigna las "
      "pide detalladas y porque demandan trabajo propio"),
     ("Guía de autores", "Leída completa el 02/09/2026 en las tres páginas de INASS. El "
      "análisis crítico está en la hoja 4, con citas literales"),
     ("Reproducible", "scripts/articulo/generar_mapeo_estructura.py"),
    ]:
        ws.append([a, b])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=9)
    estilar(ws, [26, 104])
    wb.move_sheet("0. COMO SE HIZO", offset=-5)

    wb.save(SALIDA)
    print("Generado:", SALIDA)
    print(f"  artículos: {len(ART)} · semilla IJIES: "
          f"{sum(1 for a in ART if a[0]=='IJIES')}")
    print(f"  nombres literales distintos: {len(lit)} · familias: {len(fam)}")
    for h in wb.sheetnames:
        print("   ·", h)


if __name__ == "__main__":
    main()
